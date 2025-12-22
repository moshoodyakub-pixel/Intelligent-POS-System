"""
Reports and Analytics API endpoints for the POS system.
Provides sales reports, inventory alerts, and dashboard statistics.
Includes PDF and Excel export functionality.
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from typing import Optional
from datetime import datetime, timedelta
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..database import get_db
from ..models import Transaction, Product, Vendor, Customer
from ..schemas import (
    SalesReport, InventoryAlert, InventoryAlertResponse, DashboardStats
)

router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/sales", response_model=SalesReport)
def get_sales_report(
    days: int = Query(30, ge=1, le=365, description="Number of days for the report"),
    vendor_id: Optional[int] = Query(None, description="Filter by vendor ID"),
    db: Session = Depends(get_db)
):
    """
    Generate a comprehensive sales report.
    
    - **days**: Number of days to include in the report (default: 30)
    - **vendor_id**: Optional vendor filter
    """
    period_end = datetime.utcnow()
    period_start = period_end - timedelta(days=days)
    
    # Base query for transactions in the period
    query = db.query(Transaction).filter(
        Transaction.transaction_date >= period_start,
        Transaction.transaction_date <= period_end
    )
    
    if vendor_id:
        query = query.filter(Transaction.vendor_id == vendor_id)
    
    transactions = query.all()
    
    # Calculate totals
    total_revenue = sum(t.total_price for t in transactions)
    total_transactions = len(transactions)
    avg_transaction = total_revenue / total_transactions if total_transactions > 0 else 0
    
    # Top products by revenue
    product_sales = {}
    for t in transactions:
        product = db.query(Product).filter(Product.id == t.product_id).first()
        if product:
            if product.id not in product_sales:
                product_sales[product.id] = {
                    "product_id": product.id,
                    "product_name": product.name,
                    "total_quantity": 0,
                    "total_revenue": 0
                }
            product_sales[product.id]["total_quantity"] += t.quantity
            product_sales[product.id]["total_revenue"] += t.total_price
    
    top_products = sorted(
        product_sales.values(),
        key=lambda x: x["total_revenue"],
        reverse=True
    )[:10]
    
    # Sales by vendor
    vendor_sales = {}
    for t in transactions:
        vendor = db.query(Vendor).filter(Vendor.id == t.vendor_id).first()
        if vendor:
            if vendor.id not in vendor_sales:
                vendor_sales[vendor.id] = {
                    "vendor_id": vendor.id,
                    "vendor_name": vendor.name,
                    "total_transactions": 0,
                    "total_revenue": 0
                }
            vendor_sales[vendor.id]["total_transactions"] += 1
            vendor_sales[vendor.id]["total_revenue"] += t.total_price
    
    sales_by_vendor = sorted(
        vendor_sales.values(),
        key=lambda x: x["total_revenue"],
        reverse=True
    )
    
    # Sales trend (daily aggregation)
    daily_sales = {}
    for t in transactions:
        date_key = t.transaction_date.strftime('%Y-%m-%d')
        if date_key not in daily_sales:
            daily_sales[date_key] = {"date": date_key, "revenue": 0, "transactions": 0}
        daily_sales[date_key]["revenue"] += t.total_price
        daily_sales[date_key]["transactions"] += 1
    
    sales_trend = sorted(daily_sales.values(), key=lambda x: x["date"])
    
    return SalesReport(
        total_revenue=round(total_revenue, 2),
        total_transactions=total_transactions,
        average_transaction_value=round(avg_transaction, 2),
        top_products=top_products,
        sales_by_vendor=sales_by_vendor,
        sales_trend=sales_trend,
        period_start=period_start,
        period_end=period_end
    )


@router.get("/inventory-alerts", response_model=InventoryAlertResponse)
def get_inventory_alerts(
    critical_threshold: int = Query(5, ge=0, description="Critical stock level"),
    warning_threshold: int = Query(15, ge=0, description="Warning stock level"),
    low_threshold: int = Query(25, ge=0, description="Low stock level"),
    db: Session = Depends(get_db)
):
    """
    Get inventory alerts for products with low stock.
    
    - **critical_threshold**: Stock level for critical alert (default: 5)
    - **warning_threshold**: Stock level for warning alert (default: 15)
    - **low_threshold**: Stock level for low alert (default: 25)
    """
    # Get all products with stock below low_threshold
    products = db.query(Product).filter(Product.quantity <= low_threshold).all()
    
    alerts = []
    critical_count = 0
    warning_count = 0
    low_count = 0
    
    for product in products:
        vendor = db.query(Vendor).filter(Vendor.id == product.vendor_id).first()
        vendor_name = vendor.name if vendor else "Unknown"
        
        # Determine alert level
        if product.quantity <= critical_threshold:
            alert_level = "critical"
            critical_count += 1
        elif product.quantity <= warning_threshold:
            alert_level = "warning"
            warning_count += 1
        else:
            alert_level = "low"
            low_count += 1
        
        alerts.append(InventoryAlert(
            product_id=product.id,
            product_name=product.name,
            current_quantity=product.quantity,
            threshold=low_threshold,
            vendor_id=product.vendor_id,
            vendor_name=vendor_name,
            alert_level=alert_level
        ))
    
    # Sort by alert level (critical first)
    level_order = {"critical": 0, "warning": 1, "low": 2}
    alerts.sort(key=lambda x: (level_order[x.alert_level], x.current_quantity))
    
    return InventoryAlertResponse(
        alerts=alerts,
        total_alerts=len(alerts),
        critical_count=critical_count,
        warning_count=warning_count,
        low_count=low_count
    )


@router.get("/dashboard-stats", response_model=DashboardStats)
def get_dashboard_stats(
    days: int = Query(7, ge=1, le=30, description="Days for trend data"),
    db: Session = Depends(get_db)
):
    """
    Get dashboard statistics including totals and trends.
    
    - **days**: Number of days for revenue trend (default: 7)
    """
    # Total counts
    total_products = db.query(Product).count()
    total_vendors = db.query(Vendor).count()
    total_transactions = db.query(Transaction).count()
    
    # Total revenue
    total_revenue_result = db.query(func.sum(Transaction.total_price)).scalar()
    total_revenue = float(total_revenue_result) if total_revenue_result else 0
    
    # Low stock count (products with quantity <= 10)
    low_stock_count = db.query(Product).filter(Product.quantity <= 10).count()
    
    # Recent transactions
    recent = db.query(Transaction).order_by(
        desc(Transaction.transaction_date)
    ).limit(5).all()
    
    recent_transactions = []
    for t in recent:
        product = db.query(Product).filter(Product.id == t.product_id).first()
        vendor = db.query(Vendor).filter(Vendor.id == t.vendor_id).first()
        recent_transactions.append({
            "id": t.id,
            "product_name": product.name if product else "Unknown",
            "vendor_name": vendor.name if vendor else "Unknown",
            "quantity": t.quantity,
            "total_price": t.total_price,
            "date": t.transaction_date.isoformat()
        })
    
    # Revenue trend
    period_start = datetime.utcnow() - timedelta(days=days)
    transactions = db.query(Transaction).filter(
        Transaction.transaction_date >= period_start
    ).all()
    
    daily_revenue = {}
    for t in transactions:
        date_key = t.transaction_date.strftime('%Y-%m-%d')
        if date_key not in daily_revenue:
            daily_revenue[date_key] = 0
        daily_revenue[date_key] += t.total_price
    
    revenue_trend = [
        {"date": k, "revenue": round(v, 2)}
        for k, v in sorted(daily_revenue.items())
    ]
    
    return DashboardStats(
        total_products=total_products,
        total_vendors=total_vendors,
        total_transactions=total_transactions,
        total_revenue=round(total_revenue, 2),
        low_stock_count=low_stock_count,
        recent_transactions=recent_transactions,
        revenue_trend=revenue_trend
    )


@router.get("/analytics/product/{product_id}")
def get_product_analytics(
    product_id: int,
    days: int = Query(30, ge=1, le=365, description="Days for analysis"),
    db: Session = Depends(get_db)
):
    """
    Get detailed analytics for a specific product.
    
    - **product_id**: Product ID to analyze
    - **days**: Number of days for the analysis (default: 30)
    """
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    period_start = datetime.utcnow() - timedelta(days=days)
    
    transactions = db.query(Transaction).filter(
        Transaction.product_id == product_id,
        Transaction.transaction_date >= period_start
    ).all()
    
    total_sold = sum(t.quantity for t in transactions)
    total_revenue = sum(t.total_price for t in transactions)
    transaction_count = len(transactions)
    
    # Daily sales
    daily_sales = {}
    for t in transactions:
        date_key = t.transaction_date.strftime('%Y-%m-%d')
        if date_key not in daily_sales:
            daily_sales[date_key] = {"quantity": 0, "revenue": 0}
        daily_sales[date_key]["quantity"] += t.quantity
        daily_sales[date_key]["revenue"] += t.total_price
    
    sales_trend = [
        {"date": k, **v}
        for k, v in sorted(daily_sales.items())
    ]
    
    # Average daily sales
    avg_daily_quantity = total_sold / days if days > 0 else 0
    avg_daily_revenue = total_revenue / days if days > 0 else 0
    
    # Estimated days until stock runs out
    days_until_stockout = (
        product.quantity / avg_daily_quantity 
        if avg_daily_quantity > 0 else float('inf')
    )
    
    return {
        "product": {
            "id": product.id,
            "name": product.name,
            "current_stock": product.quantity,
            "price": product.price
        },
        "period": {
            "start": period_start.isoformat(),
            "end": datetime.utcnow().isoformat(),
            "days": days
        },
        "sales_summary": {
            "total_quantity_sold": total_sold,
            "total_revenue": round(total_revenue, 2),
            "transaction_count": transaction_count,
            "average_daily_quantity": round(avg_daily_quantity, 2),
            "average_daily_revenue": round(avg_daily_revenue, 2)
        },
        "stock_forecast": {
            "days_until_stockout": round(days_until_stockout, 1) if days_until_stockout != float('inf') else None,
            "reorder_recommended": product.quantity <= 10 or (days_until_stockout < 7 and days_until_stockout != float('inf'))
        },
        "sales_trend": sales_trend
    }


def _generate_sales_report_data(db: Session, days: int, vendor_id: Optional[int] = None):
    """Helper function to generate sales report data for export."""
    period_end = datetime.utcnow()
    period_start = period_end - timedelta(days=days)
    
    query = db.query(Transaction).filter(
        Transaction.transaction_date >= period_start,
        Transaction.transaction_date <= period_end
    )
    
    if vendor_id:
        query = query.filter(Transaction.vendor_id == vendor_id)
    
    transactions = query.all()
    
    total_revenue = sum(t.total_price for t in transactions)
    total_transactions = len(transactions)
    avg_transaction = total_revenue / total_transactions if total_transactions > 0 else 0
    
    # Top products
    product_sales = {}
    for t in transactions:
        product = db.query(Product).filter(Product.id == t.product_id).first()
        if product:
            if product.id not in product_sales:
                product_sales[product.id] = {
                    "product_name": product.name,
                    "total_quantity": 0,
                    "total_revenue": 0
                }
            product_sales[product.id]["total_quantity"] += t.quantity
            product_sales[product.id]["total_revenue"] += t.total_price
    
    top_products = sorted(
        product_sales.values(),
        key=lambda x: x["total_revenue"],
        reverse=True
    )[:10]
    
    return {
        "period_start": period_start,
        "period_end": period_end,
        "total_revenue": round(total_revenue, 2),
        "total_transactions": total_transactions,
        "avg_transaction": round(avg_transaction, 2),
        "top_products": top_products,
        "transactions": transactions
    }


@router.get("/export/sales/pdf")
def export_sales_report_pdf(
    days: int = Query(30, ge=1, le=365, description="Number of days for the report"),
    vendor_id: Optional[int] = Query(None, description="Filter by vendor ID"),
    db: Session = Depends(get_db)
):
    """
    Export sales report as PDF.
    
    - **days**: Number of days to include in the report (default: 30)
    - **vendor_id**: Optional vendor filter
    """
    data = _generate_sales_report_data(db, days, vendor_id)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=1  # Center
    )
    
    # Title
    elements.append(Paragraph("Sales Report", title_style))
    elements.append(Paragraph(
        f"Period: {data['period_start'].strftime('%Y-%m-%d')} to {data['period_end'].strftime('%Y-%m-%d')}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 20))
    
    # Summary Table
    summary_data = [
        ["Metric", "Value"],
        ["Total Revenue", f"${data['total_revenue']:,.2f}"],
        ["Total Transactions", str(data['total_transactions'])],
        ["Average Transaction", f"${data['avg_transaction']:,.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6'))
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Top Products Table
    if data['top_products']:
        elements.append(Paragraph("Top Selling Products", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        products_data = [["Product Name", "Quantity Sold", "Revenue"]]
        for p in data['top_products']:
            products_data.append([
                p['product_name'],
                str(p['total_quantity']),
                f"${p['total_revenue']:,.2f}"
            ])
        
        products_table = Table(products_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        products_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
        ]))
        elements.append(products_table)
    
    # Footer
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(
        f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')} | Intelligent POS System",
        ParagraphStyle('Footer', fontSize=8, textColor=colors.grey)
    ))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"sales_report_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/sales/excel")
def export_sales_report_excel(
    days: int = Query(30, ge=1, le=365, description="Number of days for the report"),
    vendor_id: Optional[int] = Query(None, description="Filter by vendor ID"),
    db: Session = Depends(get_db)
):
    """
    Export sales report as Excel file.
    
    - **days**: Number of days to include in the report (default: 30)
    - **vendor_id**: Optional vendor filter
    """
    data = _generate_sales_report_data(db, days, vendor_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "Sales Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Period: {data['period_start'].strftime('%Y-%m-%d')} to {data['period_end'].strftime('%Y-%m-%d')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Summary section
    ws['A4'] = "Summary"
    ws['A4'].font = Font(bold=True, size=12)
    
    summary_headers = ['Metric', 'Value']
    summary_data = [
        ['Total Revenue', f"${data['total_revenue']:,.2f}"],
        ['Total Transactions', data['total_transactions']],
        ['Average Transaction', f"${data['avg_transaction']:,.2f}"]
    ]
    
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, row_data in enumerate(summary_data, 6):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
    
    # Top Products section
    ws['A10'] = "Top Selling Products"
    ws['A10'].font = Font(bold=True, size=12)
    
    product_headers = ['Product Name', 'Quantity Sold', 'Revenue']
    for col, header in enumerate(product_headers, 1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, product in enumerate(data['top_products'], 12):
        ws.cell(row=row_idx, column=1, value=product['product_name']).border = border
        ws.cell(row=row_idx, column=2, value=product['total_quantity']).border = border
        ws.cell(row=row_idx, column=3, value=f"${product['total_revenue']:,.2f}").border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"sales_report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/inventory/pdf")
def export_inventory_alerts_pdf(
    critical_threshold: int = Query(5, ge=0, description="Critical stock level"),
    warning_threshold: int = Query(15, ge=0, description="Warning stock level"),
    low_threshold: int = Query(25, ge=0, description="Low stock level"),
    db: Session = Depends(get_db)
):
    """Export inventory alerts as PDF."""
    # Get products with low stock
    products = db.query(Product).filter(Product.quantity <= low_threshold).all()
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=1
    )
    
    elements.append(Paragraph("Inventory Alerts Report", title_style))
    elements.append(Paragraph(
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 20))
    
    if products:
        data = [["Product", "Current Stock", "Vendor", "Alert Level"]]
        for p in products:
            vendor = db.query(Vendor).filter(Vendor.id == p.vendor_id).first()
            if p.quantity <= critical_threshold:
                level = "CRITICAL"
            elif p.quantity <= warning_threshold:
                level = "WARNING"
            else:
                level = "LOW"
            data.append([p.name, str(p.quantity), vendor.name if vendor else "N/A", level])
        
        table = Table(data, colWidths=[2.5*inch, 1.5*inch, 2*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6'))
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("✓ All products are well stocked!", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"inventory_alerts_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/inventory/excel")
def export_inventory_alerts_excel(
    critical_threshold: int = Query(5, ge=0, description="Critical stock level"),
    warning_threshold: int = Query(15, ge=0, description="Warning stock level"),
    low_threshold: int = Query(25, ge=0, description="Low stock level"),
    db: Session = Depends(get_db)
):
    """Export inventory alerts as Excel file."""
    products = db.query(Product).filter(Product.quantity <= low_threshold).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Alerts"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    critical_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
    warning_fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
    
    ws.merge_cells('A1:D1')
    ws['A1'] = "Inventory Alerts Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Product', 'Current Stock', 'Vendor', 'Alert Level']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row_idx, p in enumerate(products, 4):
        vendor = db.query(Vendor).filter(Vendor.id == p.vendor_id).first()
        if p.quantity <= critical_threshold:
            level = "CRITICAL"
            fill = critical_fill
        elif p.quantity <= warning_threshold:
            level = "WARNING"
            fill = warning_fill
        else:
            level = "LOW"
            fill = None
        
        ws.cell(row=row_idx, column=1, value=p.name)
        ws.cell(row=row_idx, column=2, value=p.quantity)
        ws.cell(row=row_idx, column=3, value=vendor.name if vendor else "N/A")
        level_cell = ws.cell(row=row_idx, column=4, value=level)
        if fill:
            level_cell.fill = fill
            level_cell.font = Font(bold=True, color="FFFFFF" if level == "CRITICAL" else "000000")
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"inventory_alerts_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
